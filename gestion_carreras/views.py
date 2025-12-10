
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from io import BytesIO

@login_required
def reporte_pdf(request, carrera_id):
    """Generar reporte PDF del cronograma"""
    carrera = get_object_or_404(Carrera, id=carrera_id)
    estados = EstadoCronograma.objects.filter(carrera=carrera).select_related('fase').order_by('fase__orden')
    
    # Crear el PDF en memoria
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)
    
    # Contenedor para los elementos del PDF
    elements = []
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#2c3e50'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    
    # Título
    elements.append(Paragraph("Universidad Autónoma Tomás Frías", title_style))
    elements.append(Paragraph("Reporte de Rediseño Curricular", styles['Heading2']))
    elements.append(Spacer(1, 12))
    
    # Información de la carrera
    carrera_info = [
        ['Carrera:', carrera.nombre],
        ['Facultad:', carrera.facultad.nombre],
        ['Sede:', carrera.sede.nombre],
        ['Grado:', carrera.get_grado_academico_display()],
        ['Fecha:', timezone.now().strftime('%d/%m/%Y %H:%M')]
    ]
    
    info_table = Table(carrera_info, colWidths=[2*inch, 4*inch])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#ecf0f1')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    elements.append(info_table)
    elements.append(Spacer(1, 20))
    
    # Tabla de cronograma
    data = [['No.', 'Fase', 'Estado', 'Fecha Inicio', 'Fecha Conclusión']]
    
    for idx, estado in enumerate(estados, 1):
        fecha_inicio = estado.fecha_inicio.strftime('%d/%m/%Y') if estado.fecha_inicio else '-'
        fecha_conclusion = estado.fecha_conclusion.strftime('%d/%m/%Y') if estado.fecha_conclusion else '-'
        
        data.append([
            str(idx),
            f"{estado.fase.codigo}\n{estado.fase.nombre[:50]}",
            estado.get_estado_display(),
            fecha_inicio,
            fecha_conclusion
        ])
    
    # Crear tabla
    table = Table(data, colWidths=[0.5*inch, 3*inch, 1.2*inch, 1.2*inch, 1.2*inch])
    
    # Estilo de la tabla
    table_style = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]
    
    # Colorear filas según estado
    for idx, estado in enumerate(estados, 1):
        if estado.estado == 'completado':
            table_style.append(('BACKGROUND', (0, idx), (-1, idx), colors.HexColor('#e8f5e8')))
        elif estado.estado == 'en_proceso':
            table_style.append(('BACKGROUND', (0, idx), (-1, idx), colors.HexColor('#fff3e0')))
        else:
            table_style.append(('BACKGROUND', (0, idx), (-1, idx), colors.HexColor('#ffebee')))
    
    table.setStyle(TableStyle(table_style))
    elements.append(table)
    
    # Generar PDF
    doc.build(elements)
    
    # Obtener el valor del buffer y crear la respuesta
    pdf = buffer.getvalue()
    buffer.close()
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Reporte_{carrera.nombre.replace(" ", "_")}.pdf"'
    response.write(pdf)
    
    return response

@login_required
def reporte_excel(request, carrera_id):
    """Generar reporte Excel del cronograma"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    carrera = get_object_or_404(Carrera, id=carrera_id)
    estados = EstadoCronograma.objects.filter(carrera=carrera).select_related('fase').order_by('fase__orden')
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Cronograma"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="34495e", end_color="34495e", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    ws['A1'] = "Universidad Autónoma Tomás Frías"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:F1')
    
    ws['A2'] = "Reporte de Rediseño Curricular"
    ws['A2'].font = Font(bold=True, size=12)
    ws.merge_cells('A2:F2')
    
    # Información de la carrera
    ws['A4'] = "Carrera:"
    ws['B4'] = carrera.nombre
    ws['A5'] = "Facultad:"
    ws['B5'] = carrera.facultad.nombre
    ws['A6'] = "Sede:"
    ws['B6'] = carrera.sede.nombre
    ws['A7'] = "Fecha:"
    ws['B7'] = timezone.now().strftime('%d/%m/%Y %H:%M')
    
    # Encabezados de tabla
    headers = ['No.', 'Código', 'Fase/Actividad', 'Estado', 'Fecha Inicio', 'Fecha Conclusión']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=9, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Datos
    for idx, estado in enumerate(estados, 1):
        row = idx + 9
        
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=estado.fase.codigo)
        ws.cell(row=row, column=3, value=estado.fase.nombre)
        ws.cell(row=row, column=4, value=estado.get_estado_display())
        ws.cell(row=row, column=5, value=estado.fecha_inicio.strftime('%d/%m/%Y') if estado.fecha_inicio else '-')
        ws.cell(row=row, column=6, value=estado.fecha_conclusion.strftime('%d/%m/%Y') if estado.fecha_conclusion else '-')
        
        # Aplicar bordes y colores
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Color según estado
            if estado.estado == 'completado':
                cell.fill = PatternFill(start_color="e8f5e8", end_color="e8f5e8", fill_type="solid")
            elif estado.estado == 'en_proceso':
                cell.fill = PatternFill(start_color="fff3e0", end_color="fff3e0", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="ffebee", end_color="ffebee", fill_type="solid")
    
    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    
    # Guardar en memoria
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="Reporte_{carrera.nombre.replace(" ", "_")}.xlsx"'
    
    return response