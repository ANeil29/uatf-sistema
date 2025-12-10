
import io
from datetime import datetime
from django.http import HttpResponse
from django.template.loader import render_to_string
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from .models import Carrera, EstadoCronograma

def generar_reporte_pdf(carrera_id):
    """Genera reporte PDF detallado de una carrera"""
    carrera = Carrera.objects.get(id=carrera_id)
    estados = EstadoCronograma.objects.filter(carrera=carrera).select_related('fase')
    
    # Calcular progreso
    total_fases = estados.count()
    fases_completadas = estados.filter(estado='completado').count()
    progreso_porcentaje = (fases_completadas / total_fases * 100) if total_fases > 0 else 0
    
    # Crear buffer para el PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1*inch)
    elements = []
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        alignment=1  # Centrado
    )
    
    # Título
    title = Paragraph(f"REPORTE DE PROGRESO - REDISEÑO CURRICULAR", title_style)
    elements.append(title)
    
    # Información de la carrera
    info_carrera = [
        [Paragraph('<b>Carrera:</b>', styles['Normal']), Paragraph(carrera.nombre, styles['Normal'])],
        [Paragraph('<b>Facultad:</b>', styles['Normal']), Paragraph(carrera.facultad.nombre, styles['Normal'])],
        [Paragraph('<b>Sede:</b>', styles['Normal']), Paragraph(carrera.sede.nombre, styles['Normal'])],
        [Paragraph('<b>Fecha de Reporte:</b>', styles['Normal']), Paragraph(datetime.now().strftime('%d/%m/%Y %H:%M'), styles['Normal'])],
        [Paragraph('<b>Progreso General:</b>', styles['Normal']), Paragraph(f'{progreso_porcentaje:.1f}% ({fases_completadas}/{total_fases} fases)', styles['Normal'])],
    ]
    
    info_table = Table(info_carrera, colWidths=[2*inch, 4*inch])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('PADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Tabla de estados del cronograma
    datos_estados = []
    # Encabezados
    datos_estados.append([
        Paragraph('<b>No.</b>', styles['Normal']),
        Paragraph('<b>Fase</b>', styles['Normal']),
        Paragraph('<b>Estado</b>', styles['Normal']),
        Paragraph('<b>Fechas</b>', styles['Normal']),
        Paragraph('<b>Medios Verificación</b>', styles['Normal']),
    ])
    
    # Datos
    for i, estado in enumerate(estados, 1):
        # Determinar color según estado
        if estado.estado == 'completado':
            color_fondo = colors.lightgreen
        elif estado.estado == 'en_proceso':
            color_fondo = colors.lightyellow
        else:
            color_fondo = colors.lightcoral
        
        fechas = ""
        if estado.fecha_inicio:
            fechas += f"Inicio: {estado.fecha_inicio.strftime('%d/%m/%Y')}\n"
        if estado.fecha_conclusion:
            fechas += f"Conclusión: {estado.fecha_conclusion.strftime('%d/%m/%Y')}"
        
        medios = estado.medios_verificacion or "Sin especificar"
        if len(medios) > 100:  # Limitar longitud para PDF
            medios = medios[:100] + "..."
        
        datos_estados.append([
            Paragraph(str(i), styles['Normal']),
            Paragraph(f"<b>{estado.fase.codigo}</b><br/>{estado.fase.nombre}", styles['Normal']),
            Paragraph(estado.get_estado_display(), styles['Normal']),
            Paragraph(fechas, styles['Normal']),
            Paragraph(medios, styles['Normal']),
        ])
    
    estados_table = Table(datos_estados, colWidths=[0.5*inch, 2.5*inch, 1.2*inch, 1.5*inch, 2.3*inch])
    estados_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    
    # Aplicar colores a las filas según estado
    for i in range(1, len(datos_estados)):
        estado_actual = estados[i-1].estado
        if estado_actual == 'completado':
            estados_table.setStyle(TableStyle([
                ('BACKGROUND', (0, i), (-1, i), colors.lightgreen),
            ]))
        elif estado_actual == 'en_proceso':
            estados_table.setStyle(TableStyle([
                ('BACKGROUND', (0, i), (-1, i), colors.lightyellow),
            ]))
    
    elements.append(estados_table)
    
    # Generar PDF
    doc.build(elements)
    buffer.seek(0)
    
    return buffer

def generar_reporte_excel(carrera_id):
    """Genera reporte Excel detallado de una carrera"""
    carrera = Carrera.objects.get(id=carrera_id)
    estados = EstadoCronograma.objects.filter(carrera=carrera).select_related('fase')
    
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Progreso Rediseño Curricular"
    
    # Estilos
    title_font = Font(size=16, bold=True)
    header_font = Font(size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Título
    ws.merge_cells('A1:E1')
    ws['A1'] = f"REPORTE DE PROGRESO - REDISEÑO CURRICULAR - {carrera.nombre}"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Información de la carrera
    ws['A3'] = "Carrera:"
    ws['B3'] = carrera.nombre
    ws['A4'] = "Facultad:"
    ws['B4'] = carrera.facultad.nombre
    ws['A5'] = "Sede:"
    ws['B5'] = carrera.sede.nombre
    ws['A6'] = "Fecha de Reporte:"
    ws['B6'] = datetime.now().strftime('%d/%m/%Y %H:%M')
    
    # Encabezados de la tabla
    headers = ['No.', 'Código Fase', 'Nombre Fase', 'Estado', 'Fecha Inicio', 
               'Fecha Conclusión', 'Medios Verificación', 'Observaciones']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=8, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Datos
    for row, estado in enumerate(estados, 9):
        ws.cell(row=row, column=1, value=row-8)
        ws.cell(row=row, column=2, value=estado.fase.codigo)
        ws.cell(row=row, column=3, value=estado.fase.nombre)
        ws.cell(row=row, column=4, value=estado.get_estado_display())
        ws.cell(row=row, column=5, value=estado.fecha_inicio.strftime('%d/%m/%Y') if estado.fecha_inicio else "")
        ws.cell(row=row, column=6, value=estado.fecha_conclusion.strftime('%d/%m/%Y') if estado.fecha_conclusion else "")
        ws.cell(row=row, column=7, value=estado.medios_verificacion or "")
        ws.cell(row=row, column=8, value=estado.observaciones or "")
    
    # Ajustar anchos de columna
    column_widths = [5, 12, 40, 15, 12, 12, 40, 30]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Guardar en buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer

def reporte_pdf_carrera(request, carrera_id):
    """Vista para descargar reporte PDF"""
    buffer = generar_reporte_pdf(carrera_id)
    
    response = HttpResponse(buffer, content_type='application/pdf')
    carrera = Carrera.objects.get(id=carrera_id)
    filename = f"reporte_{carrera.nombre.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response

def reporte_excel_carrera(request, carrera_id):
    """Vista para descargar reporte Excel"""
    buffer = generar_reporte_excel(carrera_id)
    
    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    carrera = Carrera.objects.get(id=carrera_id)
    filename = f"reporte_{carrera.nombre.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response